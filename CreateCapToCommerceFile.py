import os

import openpyxl

FILE_PATH = os.path.join('file', 'capToCommerce', '')

wb = openpyxl.load_workbook(FILE_PATH + 'data.xlsx', data_only=True)

firstSheetName = wb.sheetnames[0]
t_sheet = wb[firstSheetName]

additonal_dictionary = {}
# TODO 今のつくりだと一度ファイルを消さないといけない。
with open("capToCommerce.txt", mode="a", encoding='utf-8') as fileobj:
    for i in range(3, 7):  #

        # if _requestNo is None:
        #     break
        _requestNo = t_sheet.cell(row=i, column=1).value

        if _requestNo is None:
            _requestNo = ''

        _requestNo = _requestNo.rjust(10, ' ')

        _voucherDateYearMonth = t_sheet.cell(row=i, column=2).value.zfill(6)
        _voucherDateDay = t_sheet.cell(row=i, column=3).value.rjust(2, '0')
        _transferDestinationWarehouseCode = t_sheet.cell(row=i, column=4).value.rjust(3, ' ')
        _transferDestinationInventorySectionCode = t_sheet.cell(row=i, column=5).value.rjust(3, ' ')
        productCode = t_sheet.cell(row=i, column=6).value.rjust(13, ' ')
        partnerKind = t_sheet.cell(row=i, column=7).value.rjust(1, ' ')
        corporation = t_sheet.cell(row=i, column=8).value.rjust(7, ' ')
        branch = t_sheet.cell(row=i, column=9).value.rjust(4, ' ')
        sellingFloor = t_sheet.cell(row=i, column=10).value.rjust(1, ' ')
        quotaQuantity = t_sheet.cell(row=i, column=11).value.zfill(7)
        inventoryCondition = t_sheet.cell(row=i, column=12).value.rjust(1, ' ')
        lineStr = _requestNo + _voucherDateYearMonth + _voucherDateDay + _transferDestinationWarehouseCode + _transferDestinationInventorySectionCode + productCode + partnerKind + corporation + branch + sellingFloor + quotaQuantity + inventoryCondition
        lineStr += "\n"

        fileobj.write(lineStr)

    fileobj.close()
