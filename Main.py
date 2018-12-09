import os

import re
import openpyxl

from entityDictionary import EntityDictionary

e_dctionary = EntityDictionary.importER()

FILE_PATH = os.path.join('file', 'import', '')
# "file\\import\\"
INPUT_FILE_NAME = "出荷指示取消IF.xlsx"

# data_only=Trueにする？
wb = openpyxl.load_workbook(FILE_PATH + INPUT_FILE_NAME, data_only=True)

firstSheetName = wb.sheetnames[0]

t_sheet = wb[firstSheetName]
# g = t_sheet.iter_rows(min_row=2, max_row=4, min_col=1, max_col=3)
# pprint.pprint(list(g))
# 読込みバージョン
# @InputFixedLengthColumn(start = 1, end = 14)
# public String lmsShipmentAchievementNumber;
# 書込みバージョン
# @OutputFixedColumn(byteSize = 12, mode = Mode.A)
# public String businessManagementNumber;

for i in range(11, 100):  # lengthはあとから
    jp_name = t_sheet.cell(row=i, column=4).value
    attr_mode = t_sheet.cell(row=i, column=8).value
    str_position = str(t_sheet.cell(row=i, column=11).value)
    end_position = str(t_sheet.cell(row=i, column=12).value)
    # TODO Noneになったら終了、なんか式間違っているらしい
    if jp_name == None:
        break

    en_name = e_dctionary.get(jp_name, 'NonMatchMethodName')
    method_name = re.sub("_(.)", lambda x: x.group(1).upper(), en_name)

    print("/** " + jp_name + " */")
    print("@InputFixedLengthColumn(start = " + str_position + ", end = " + end_position + ")")
    print("public String " + method_name + ";")

# ここまでは共通


# for cell_obj in list(t_sheet.columns)[3]:
#    print(cell_obj.value)
