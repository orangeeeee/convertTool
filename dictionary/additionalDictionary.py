import os
import csv

import openpyxl


class AdditionalDictionary:
    # col1 = TABLE_CATALOG, col2 = TABLE_SCHEMA, col3 = TABLE_NAME, col4 = COLUMN_NAME, col4= LOGICAL_NAME
    FILE_PATH = os.path.join('file', 'dictionary', '')
    IMPORT_FILE_NAME = 'a5m2_COLUMNS'
    COLUMN_ENTITY_NAME = 'COLUMN_NAME'
    COLUMN_ENTITY_JP_NAME = 'LOGICAL_NAME'

    def __init__(self):
        return

    @classmethod
    def importDict2(cls, e_dictionary):

        FILE_PATH = os.path.join('file', 'dictionary', '')

        wb = openpyxl.load_workbook(FILE_PATH + 'additionalDictionary.xlsx', data_only=True)

        firstSheetName = wb.sheetnames[1]
        t_sheet = wb[firstSheetName]

        additonal_dictionary = {}

        for i in range(3, 1000):  # TODO lengthはあとから

            json_prop = t_sheet.cell(row=i, column=1).value

            if json_prop is None:
                break
            _jp_name = t_sheet.cell(row=i, column=2).value
            _en_name = t_sheet.cell(row=i, column=3).value

            if _jp_name not in e_dictionary:
                additonal_dictionary[_jp_name] = _en_name

        return additonal_dictionary
